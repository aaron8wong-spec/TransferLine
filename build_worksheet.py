"""
Generates worksheet.xlsx — the manual data-entry workbook for the
3-school x 3-major proof of concept.

Run once to produce the workbook:
    python build_worksheet.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF8C4")       # yellow = fill this in
EXAMPLE_FILL = PatternFill("solid", fgColor="DCEBE1")     # green = filled-in example
LOCKED_FILL = PatternFill("solid", fgColor="EAEDF2")      # gray = reference, don't edit
THIN = Side(style="thin", color="C7CCDA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SLOTS = [
    ("calc1",    "Calculus I",                      "math"),
    ("calc2",    "Calculus II",                     "math"),
    ("calc3",    "Calculus III (Multivariable)",    "math"),
    ("linalg",   "Linear Algebra",                  "math"),
    ("diffeq",   "Differential Equations",          "math"),
    ("discrete", "Discrete Mathematics",            "math"),
    ("prog1",    "Programming Methodology",         "programming"),
    ("prog2",    "Data Structures",                 "programming"),
    ("comporg",  "Computer Organization",           "programming"),
    ("digital_logic", "Digital/Logic Design",       "programming"),
    ("physics1", "Physics I \u2014 Mechanics (w/ Calc)", "science"),
    ("physics2", "Physics II \u2014 E&M (w/ Calc)",      "science"),
    ("physics3", "Physics III \u2014 Waves, Optics & Modern Physics (w/ Calc)", "science"),
    ("circuits", "Circuit/Network Analysis",        "engineering"),
    ("stats_gen","Introductory Statistics (general)","cogsci"),
    ("linguistics", "Introduction to Linguistics",  "cogsci"),
    ("psych_intro", "Introduction to Psychology",   "cogsci"),
    ("neuro_intro", "Introduction to Neuroscience/Biopsychology", "cogsci"),
    ("cogsci_research", "Introduction to Research (Methods)", "cogsci"),
]

# All 4 school-major combos are now CONFIRMED via real ASSIST agreement PDFs
# (both CCCs, uploaded by the user), not general knowledge or guesses.
SCHOOL_MAJORS = [
    ("ucsd", "UC San Diego", "compe", "Computer Engineering, B.S. (CSE Dept.)",
     "CONFIRMED via actual UCSD ASSIST agreements (both De Anza and EVC PDFs agree). Official 'General "
     "Advice' list: MATH 20A/20B/20C (calc1/2/3), MATH 20D (diffeq), MATH 18 (linalg), calc-based physics "
     "series (physics1/2/3), CSE 8B-or-11 (prog1), CSE 12 (prog2), CSE 20 (discrete). CORRECTION from an "
     "earlier version of this sheet: discrete math IS explicitly required (previously left unmarked); "
     "Computer Organization is NOT in the official required list (previously marked required in error) -- "
     "it appears in the detailed per-CCC table but not the summary bullet list, so treat it as unconfirmed."),

    ("ucsd", "UC San Diego", "cogsci", "Cognitive Science, B.S.",
     "CONFIRMED via actual UCSD ASSIST agreements (both CCCs). Students choose ONE math sequence: "
     "MATH 10A/10B/10C + MATH 18, OR MATH 20A/20B + MATH 18 -- represented here using the existing "
     "calc1/calc2/linalg slots (the 20AB+18 path), since that's the complete path both CCCs can satisfy. "
     "cogsci_research (PSYC 2/PSYCH 018 + a stats course) is ALSO mandatory -- user confirmed via the "
     "actual ASSIST page after this sheet initially miscategorized it as elective (the PDF-to-text "
     "extraction made the elective-vs-required boundary ambiguous; the live page resolved it). Beyond "
     "that, a genuinely elective 'pick 1 from each of 2 remaining groups' covers neuro electives and "
     "an alternate intro-programming pool -- see Elective Groups/Options tabs."),

    ("uci", "UC Irvine", "compe", "Computer Engineering (Henry Samueli School of Engineering)",
     "CONFIRMED via actual UCI ASSIST agreements (both CCCs, identical requirement list). UCI explicitly "
     "separates 'required for admission' from 'recommended for admission/time to degree' -- ONLY the "
     "former is marked required here: calc1/2/3 (MATH 2A/2B/2D), diffeq (MATH 3D), linalg (MATH 3A), "
     "physics1/2/3, prog1 (Programming Concepts I), and circuits (Circuit Analysis). Programming II, "
     "Computer Organization, and Digital Logic are explicitly 'recommended, not required for admission' -- "
     "left unmarked here on purpose, not an oversight."),

    ("uci", "UC Irvine", "cogsci", "Cognitive Sciences (School of Social Sciences)",
     "CONFIRMED via actual ASSIST agreement PDFs for BOTH CCCs (2025-2026). calc1 (MATH 2A) and calc2 "
     "(MATH 2B) required; COGS 9A/9C need CCC equivalents, COGS 9B is UCI-only (no CCC option exists, "
     "ever -- not missing data). Beyond that, a 'complete 3 semesters/6 courses' elective requirement "
     "spans math extensions, programming, physics, linguistics, neuro electives, or logic -- see 'Elective "
     "Groups'/'Elective Options' tabs, now populated for BOTH CCCs."),
]

CCCS = ["deanza", "evc"]
CCC_LABELS = {
    "deanza": "De Anza College", "evc": "Evergreen Valley College",
}

# Slots CONFIRMED via real ASSIST agreement PDFs this session (not guesses) --
# shown in green regardless of whether the whole row is confirmed or just
# part of it.
CONFIRMED_MARKS = {
    ("ucsd", "compe"): {"calc1", "calc2", "calc3", "diffeq", "linalg",
                        "prog1", "prog2", "discrete",
                        "physics1", "physics2", "physics3"},
    ("ucsd", "cogsci"): {"calc1", "calc2", "linalg", "cogsci_research"},
    ("uci", "compe"): {"calc1", "calc2", "calc3", "diffeq", "linalg",
                       "prog1", "circuits",
                       "physics1", "physics2", "physics3"},
    ("uci", "cogsci"): {"calc1", "calc2", "psych_intro"},  # unconditionally required part of UCI's list
}


# Groups of "pick N of these M options" requirements -- for when a major's
# requirement genuinely can't be flattened into a fixed list of required
# slots (see UCI Cognitive Sciences below). Each option can itself bundle
# multiple slots together (AND within an option), and a student satisfies
# the group by completing pick_n whole options.
#
# This example is now EVC-SPECIFIC real data, extracted from an actual ASSIST
# agreement PDF (EVC -> UCI Cognitive Sciences, 2025-2026, published 6/10/26)
# the user uploaded. UCI's requirement is "3 semesters/quarters, 6 courses
# recommended" spread across math extensions / programming / physics /
# linguistics / neuro electives / logic -- EVC does NOT offer linguistics or
# extra neuro elective equivalents at all, so those aren't listed as usable
# options here. De Anza's available options will likely differ and need
# their own PDF to confirm.
ELECTIVE_GROUPS = [
    # group_id, school_key, major_id, pick_n, description
    ("uci_cogsci_electives_evc", "uci", "cogsci", "enough options to reach 6 courses total (verify exact rule against source PDF)",
     "EVC->UCI Cognitive Sciences: complete 3 semesters/quarters (6 courses recommended) from the options "
     "below. CONFIRMED from actual ASSIST PDF, but the precise UCI-course-to-EVC-course row pairing was "
     "ambiguous in the extracted text for a couple of these -- flagged per-option below."),
    ("uci_cogsci_electives_deanza", "uci", "cogsci", "enough options to reach 6 courses total (verify exact rule against source PDF)",
     "De Anza->UCI Cognitive Sciences: same 'complete 3 semesters/6 courses' requirement as EVC, but De "
     "Anza has MUCH better coverage -- it offers real equivalents for linguistics and logic that EVC "
     "doesn't. WARNING: the PHYS2A/2B/2C physics option is flagged in the source PDF as 'Effective next "
     "fall, this articulation is no longer valid' -- use the PHYS4A/4B/4C option instead going forward."),
    ("ucsd_cogsci_electives_evc", "ucsd", "cogsci", "1 course from each of 2 groups (a neuro/programming pool, and intro programming)",
     "EVC->UCSD Cognitive Science: beyond calc1/calc2/linalg/cogsci_research (all now confirmed mandatory, "
     "see Requirements Matrix), students also complete 1 course from each of 2 further groups. Some "
     "UCI-course-to-CCC-course pairings were ambiguous in the extracted PDF text -- flagged per-option below."),
    ("ucsd_cogsci_electives_deanza", "ucsd", "cogsci", "1 course from each of 2 groups (a neuro/programming pool, and intro programming)",
     "De Anza->UCSD Cognitive Science: same structure as EVC. De Anza has no clean research-methods "
     "equivalent found in the source PDF (mostly 'No Course Articulated') but does offer real programming "
     "options."),
]

ELECTIVE_OPTIONS = [
    # group_id, option_id, option_label, slot_id (blank = not mapped to a tracked slot)
    ("uci_cogsci_electives_evc", "programming", "Programming series: EVC COMSC 075 + COMSC 076 (2 courses)", "prog1"),
    ("uci_cogsci_electives_evc", "programming", "Programming series: EVC COMSC 075 + COMSC 076 (2 courses)", "prog2"),
    ("uci_cogsci_electives_evc", "multivar_calc", "EVC MATH 073 (Multivariable Calculus, 5 units)", "calc3"),
    ("uci_cogsci_electives_evc", "linear_algebra", "EVC MATH 079 (Linear Algebra, 3 units)", "linalg"),
    ("uci_cogsci_electives_evc", "diffeq", "EVC MATH 078 (Differential Equations, 4 units)", "diffeq"),
    ("uci_cogsci_electives_evc", "physics_2course", "EVC PHYS 007A + PHYS 007B (only 2 courses -- EVC has no 3rd physics course option here)", "physics1"),
    ("uci_cogsci_electives_evc", "physics_2course", "EVC PHYS 007A + PHYS 007B (only 2 courses -- EVC has no 3rd physics course option here)", "physics2"),
    ("uci_cogsci_electives_evc", "statistics", "EVC STAT C1000 (or BUS 060)", "stats_gen"),
    ("uci_cogsci_electives_evc", "logic", "EVC PHIL 090 (Introduction to Logic) -- no matching slot tracked here", ""),
    ("uci_cogsci_electives_evc", "linguistics", "No EVC equivalent exists for any linguistics option (LSCI 3/10/20/51)", ""),
    ("uci_cogsci_electives_evc", "neuro_electives", "No EVC equivalent exists for extra neuro electives (BIO SCI 36/37/38)", ""),

    ("uci_cogsci_electives_deanza", "programming", "Programming series: De Anza CIS 40 + CIS 41A + CIS 41B (3 courses)", "prog1"),
    ("uci_cogsci_electives_deanza", "programming", "Programming series: De Anza CIS 40 + CIS 41A + CIS 41B (3 courses)", "prog2"),
    ("uci_cogsci_electives_deanza", "linguistics", "De Anza LING 1 (Introduction to Linguistics) -- real equivalent, unlike EVC", ""),
    ("uci_cogsci_electives_deanza", "logic", "De Anza PHIL 7 (Deductive Logic, or 7H honors) -- real equivalent, unlike EVC", ""),
    ("uci_cogsci_electives_deanza", "multivar_calc", "De Anza MATH 1D (Calculus IV) -- satisfies UCI's MATH 2D/2E multivariable pair", "calc3"),
    ("uci_cogsci_electives_deanza", "linear_algebra", "De Anza MATH 2B (or 2BH)", "linalg"),
    ("uci_cogsci_electives_deanza", "diffeq", "De Anza MATH 2A (or 2AH) -- note De Anza's OWN numbering: their 'MATH 2A' means Diff Eq, not Calc I", "diffeq"),
    ("uci_cogsci_electives_deanza", "physics_3course", "De Anza PHYS 2A/2B/2C (5 units, DEPRECATED next fall) OR PHYS 4A/4B/4C (6 units, use this one going forward)", "physics1"),
    ("uci_cogsci_electives_deanza", "physics_3course", "De Anza PHYS 2A/2B/2C (5 units, DEPRECATED next fall) OR PHYS 4A/4B/4C (6 units, use this one going forward)", "physics2"),
    ("uci_cogsci_electives_deanza", "physics_3course", "De Anza PHYS 2A/2B/2C (5 units, DEPRECATED next fall) OR PHYS 4A/4B/4C (6 units, use this one going forward)", "physics3"),
    ("uci_cogsci_electives_deanza", "statistics", "De Anza STAT C1000 (or Honors), or PSYC 15 / POLI 20 (same as SOC 15)", "stats_gen"),

    ("ucsd_cogsci_electives_evc", "neuro", "EVC PSYCH 030 (Intro Biological Psychology) -- likely satisfies a neuro-focused option", "neuro_intro"),
    ("ucsd_cogsci_electives_evc", "programming", "EVC COMSC 075, or CIT 044 (Java) + COMSC 076 combo -- ambiguous exact UCSD target course", "prog1"),

    ("ucsd_cogsci_electives_deanza", "neuro", "De Anza PSYC 24 (Introduction to Psychobiology) -- likely satisfies COGS 17 Neurobiology of Cognition", "neuro_intro"),
    ("ucsd_cogsci_electives_deanza", "programming", "De Anza: mostly 'No Course Articulated' for the COGS18/CSE6R/CSE8A/CSE11-style intro Python options; CIS 40 Python + CIS 41A, or CIS 35A/36B Java, may partially apply -- needs your visual check of the actual table", "prog1"),
]


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def build():
    wb = openpyxl.Workbook()

    # ---------------- Sheet: Instructions ----------------
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Transfer Track \u2014 manual data-entry worksheet", 16, True),
        ("", None, False),
        ("Proof of concept: 2 UC schools \u00d7 2 majors, plus a shared CCC course catalog for 2 CCCs.", 11, True),
        ("", None, False),
        ("HOW TO FILL THIS IN", 12, True),
        ("1. Open assist.org in your regular web browser (this works fine \u2014 only the automated", 10, False),
        ("   API script was blocked, the website itself is public).", 10, False),
        ("2. Pick a sending college (a CCC) and a receiving school (one of the 2 UCs below).", 10, False),
        ("3. Find the agreement for the major listed on the 'Requirements Matrix' tab.", 10, False),
        ("4. For each course ASSIST lists as required, mark an X in the matching slot column", 10, False),
        ("   on 'Requirements Matrix' (yellow cells) \u2014 this only needs doing ONCE per school+major,", 10, False),
        ("   not per CCC, since it's the university's requirement, not the CCC's.", 10, False),
        ("5. On the 'CCC Catalog' tab, fill in which course AT YOUR CCC satisfies each slot", 10, False),
        ("   (yellow cells) \u2014 this is the part that differs by CCC.", 10, False),
        ("6. SEQUENCES (must take multiple CCC courses together for one requirement): just type", 10, False),
        ("   both codes in a single 'CCC Catalog' cell joined by ' + ', e.g. 'CIS 40 + CIS 41'.", 10, False),
        ("   That's one answer, not two, so it stays in one cell.", 10, False),
        ("7. ALTERNATIVES ('pick N of these M different requirements'): do NOT try to force this", 10, False),
        ("   into X's on Requirements Matrix -- it can't express that shape. Instead use the", 10, False),
        ("   'Elective Groups' and 'Elective Options' tabs (see the worked UCI Cognitive Sciences", 10, False),
        ("   example there). Only mark a slot with X on Requirements Matrix if it's required", 10, False),
        ("   outright, with no substitution allowed.", 10, False),
        ("", None, False),
        ("COLOR KEY", 12, True),
        ("Yellow  = fill this in", 10, False),
        ("Green   = example already filled in for you, showing the expected format", 10, False),
        ("Gray    = reference only, don't edit", 10, False),
        ("", None, False),
        ("WHAT HAPPENS NEXT", 12, True),
        ("Once filled in, run convert_worksheet.py (see the file next to this one) to turn this", 10, False),
        ("into the JSON files build_dist.py already knows how to read \u2014 no changes needed to the", 10, False),
        ("rest of your pipeline.", 10, False),
        ("", None, False),
        ("KNOWN LIMITATION: convert_worksheet.py currently reads 'Elective Groups'/'Elective Options'", 10, False),
        ("for your own reference but does NOT yet feed pick-N-of-M logic into the app -- the app's", 10, False),
        ("data model only understands flat 'required' lists right now. The data is captured here so", 10, False),
        ("it isn't lost, ready for whenever the app/frontend is extended to actually display choices.", 10, False),
    ]
    for i, (text, size, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name=FONT_NAME, size=size or 10, bold=bold)

    # ---------------- Sheet: Requirements Matrix ----------------
    ws2 = wb.create_sheet("Requirements Matrix")
    ws2.sheet_view.showGridLines = False
    headers = ["school_key", "School", "major_id", "Major (official name)", "Source note"] + [s[1] for s in SLOTS] + ["Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        style_header(cell)
    ws2.freeze_panes = "F2"
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 42
    ws2.column_dimensions["E"].width = 46
    for col in range(6, 6 + len(SLOTS)):
        ws2.column_dimensions[get_column_letter(col)].width = 14
    ws2.column_dimensions[get_column_letter(6 + len(SLOTS))].width = 30

    dv = DataValidation(type="list", formula1='"X,"', allow_blank=True)
    ws2.add_data_validation(dv)

    for r, (school_key, school_name, major_id, major_label, source_note) in enumerate(SCHOOL_MAJORS, start=2):
        confirmed_slots = CONFIRMED_MARKS.get((school_key, major_id), set())
        is_example = (school_key, major_id) == ("ucsd", "compe")  # the fully-worked example row
        row_fill = EXAMPLE_FILL if is_example else None

        ws2.cell(row=r, column=1, value=school_key).font = Font(name=FONT_NAME, size=10)
        ws2.cell(row=r, column=2, value=school_name).font = Font(name=FONT_NAME, size=10, bold=True)
        ws2.cell(row=r, column=3, value=major_id).font = Font(name=FONT_NAME, size=10)
        ws2.cell(row=r, column=4, value=major_label).font = Font(name=FONT_NAME, size=10)
        note_cell = ws2.cell(row=r, column=5, value=source_note)
        note_cell.font = Font(name=FONT_NAME, size=9, italic=True)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")

        for i, (slot_id, slot_name, cat) in enumerate(SLOTS):
            col = 6 + i
            cell = ws2.cell(row=r, column=col)
            cell.border = BORDER
            if slot_id in confirmed_slots:
                cell.value = "X"
                cell.fill = EXAMPLE_FILL
            else:
                cell.fill = INPUT_FILL
                dv.add(cell)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="center")

        notes_cell = ws2.cell(row=r, column=6 + len(SLOTS))
        notes_cell.fill = EXAMPLE_FILL if is_example else INPUT_FILL
        notes_cell.border = BORDER
        notes_cell.font = Font(name=FONT_NAME, size=9)
        if is_example:
            notes_cell.value = "Fully confirmed row -- see source note. Computer Organization intentionally left blank (not in official required list)."
        elif school_key == "uci" and major_id == "cogsci":
            notes_cell.value = ("psych_intro here is a proxy for COGS 9A+9C (both need PSYC C1000 + a second "
                                 "CCC course -- see Elective Options). COGS 9B has NO CCC equivalent, ever -- "
                                 "it's UCI-only by design, not a data gap. The 'pick from A/B/C' elective piece "
                                 "is in the Elective Groups/Options tabs, now covering BOTH CCCs.")

    # ---------------- Sheet: CCC Catalog ----------------
    ws3 = wb.create_sheet("CCC Catalog")
    ws3.sheet_view.showGridLines = False
    headers3 = ["slot_id", "Slot (canonical requirement)"] + [CCC_LABELS[c] for c in CCCS] + ["Notes"]
    for col, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        style_header(cell)
    ws3.freeze_panes = "C2"
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 34
    for i in range(len(CCCS)):
        ws3.column_dimensions[get_column_letter(3 + i)].width = 18
    ws3.column_dimensions[get_column_letter(3 + len(CCCS))].width = 30

    # Confirmed-real data, sourced from actual ASSIST agreement PDFs the user
    # uploaded this session (both CCCs, all 4 majors) -- supersedes earlier,
    # less precise general-knowledge guesses.
    #
    # IMPORTANT LIMITATION surfaced by this data: De Anza's "calc3" course
    # actually DIFFERS depending on which university it's satisfying --
    # MATH 1D alone for UCI, but MATH 1C + MATH 1D combined for UCSD. The
    # CCC Catalog tab (and the app's data model) only supports ONE course
    # string per (CCC, slot) pair, with no per-university variation. The
    # value below documents both cases in one string as a stopgap; if this
    # matters for your build, the data model itself may need a
    # (ccc, slot, school) key instead of just (ccc, slot).
    evc_confirmed = {
        "calc1": "MATH 071 (or MATH 066, or MATH 062 non-STEM)",
        "calc2": "MATH 072 (or MATH 067)",
        "calc3": "MATH 073",
        "diffeq": "MATH 078",
        "linalg": "MATH 079",
        "discrete": "COMSC 080 (or MATH 070) -- confirmed via UCI PDF; UCSD PDF shows NO EVC equivalent (permanent gap for UCSD path)",
        "prog1": "COMSC 075 (or CIT 044 Java)",
        "prog2": "COMSC 076",
        "circuits": "ENGR 071 (Introduction to Circuit Analysis)",
        "physics1": "PHYS 007A (or PHYS 004A General Physics)",
        "physics2": "PHYS 007B",
        "physics3": "PHYS 007C",
        "stats_gen": "STAT C1000 (or BUS 060)",
        "cogsci_research": "PSYCH 018 (Introduction to Research Methods) + STAT C1000",
    }
    deanza_confirmed = {
        "calc1": "MATH 1A (or 1AH, or MATH 12 for the business-calc path)",
        "calc2": "MATH 1B (or 1BH)",
        "calc3": "MATH 1D for UCI (multivariable) -- but MATH 1C + MATH 1D COMBINED for UCSD (see limitation note above)",
        "diffeq": "MATH 2A (or 2AH) -- NOTE: De Anza's OWN numbering means their 'MATH 2A' is Diff Eq, not Calc I",
        "linalg": "MATH 2B (or 2BH) -- NOTE: De Anza's 'MATH 2B' is Linear Algebra, not Calc II",
        "discrete": "MATH 22 (or 22H) for UCI; NO equivalent exists for UCSD's CSE 20 (permanent gap for UCSD path)",
        "prog1": "CIS 35A (or 36B) for UCSD; CIS 22A/26A/35A/36A/36B (any of these) for UCI",
        "prog2": "CIS 22C (or 22CH)",
        "circuits": "ENGR 37 (Introduction to Circuit Analysis)",
        "physics1": "PHYS 4A (also PHYS 2A works for UCI Cogsci but is being DEPRECATED next fall)",
        "physics2": "PHYS 4B (also PHYS 2B, same deprecation warning)",
        "physics3": "PHYS 4C (also PHYS 2C, same deprecation warning)",
        "stats_gen": "STAT C1000 (or Honors), or PSYC 15 / POLI 20",
        "cogsci_research": "PSYC 2 (Research Methods in Psychology) + STAT C1000 (or PSYC 15, same as SOC 15)",
    }
    CCC_CONFIRMED = {"evc": evc_confirmed, "deanza": deanza_confirmed}

    for r, (slot_id, slot_name, cat) in enumerate(SLOTS, start=2):
        ws3.cell(row=r, column=1, value=slot_id).font = Font(name=FONT_NAME, size=10)
        ws3.cell(row=r, column=2, value=slot_name).font = Font(name=FONT_NAME, size=10, bold=True)
        any_confirmed_in_row = False
        for i, ccc in enumerate(CCCS):
            col = 3 + i
            cell = ws3.cell(row=r, column=col)
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="center")
            confirmed = CCC_CONFIRMED.get(ccc, {})
            if slot_id in confirmed:
                cell.value = confirmed[slot_id]
                cell.fill = EXAMPLE_FILL
                any_confirmed_in_row = True
            else:
                cell.fill = INPUT_FILL
        notes_cell = ws3.cell(row=r, column=3 + len(CCCS))
        notes_cell.fill = INPUT_FILL
        notes_cell.border = BORDER
        notes_cell.font = Font(name=FONT_NAME, size=9)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if any_confirmed_in_row:
            notes_cell.value = "CONFIRMED via actual ASSIST agreement PDFs (2025-2026, this session)."
        if slot_id == "calc3":
            notes_cell.value = ("SEE LIMITATION NOTE: De Anza's calc3 course differs by receiving school "
                                 "(MATH 1D alone for UCI, MATH 1C+1D combined for UCSD) -- one cell can't "
                                 "fully capture this.")

    # ---------------- Sheet: Elective Groups ----------------
    ws5 = wb.create_sheet("Elective Groups")
    ws5.sheet_view.showGridLines = False
    headers5 = ["group_id", "school_key", "major_id", "pick_n (how many options required)", "Description"]
    for col, h in enumerate(headers5, start=1):
        cell = ws5.cell(row=1, column=col, value=h)
        style_header(cell)
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["B"].width = 12
    ws5.column_dimensions["C"].width = 10
    ws5.column_dimensions["D"].width = 16
    ws5.column_dimensions["E"].width = 70

    for r, (group_id, school_key, major_id, pick_n, description) in enumerate(ELECTIVE_GROUPS, start=2):
        vals = [group_id, school_key, major_id, pick_n, description]
        for c, v in enumerate(vals, start=1):
            cell = ws5.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = EXAMPLE_FILL
            cell.border = BORDER
            if c == 5:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    # A few blank template rows for the user to add their own groups
    next_row = len(ELECTIVE_GROUPS) + 2
    for r in range(next_row, next_row + 4):
        for c in range(1, 6):
            cell = ws5.cell(row=r, column=c)
            cell.fill = INPUT_FILL
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=10)

    # ---------------- Sheet: Elective Options ----------------
    ws6 = wb.create_sheet("Elective Options")
    ws6.sheet_view.showGridLines = False
    headers6 = ["group_id", "option_id", "option_label", "slot_id (blank if untracked)"]
    for col, h in enumerate(headers6, start=1):
        cell = ws6.cell(row=1, column=col, value=h)
        style_header(cell)
    ws6.column_dimensions["A"].width = 24
    ws6.column_dimensions["B"].width = 16
    ws6.column_dimensions["C"].width = 55
    ws6.column_dimensions["D"].width = 20

    valid_slot_ids = [s[0] for s in SLOTS]
    dv_slot = DataValidation(type="list", formula1=f'"{",".join(valid_slot_ids)},"', allow_blank=True)
    ws6.add_data_validation(dv_slot)

    for r, (group_id, option_id, option_label, slot_id) in enumerate(ELECTIVE_OPTIONS, start=2):
        vals = [group_id, option_id, option_label, slot_id]
        for c, v in enumerate(vals, start=1):
            cell = ws6.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = EXAMPLE_FILL
            cell.border = BORDER
            if c == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    # A few blank template rows for the user's own options, with slot_id dropdown
    next_row = len(ELECTIVE_OPTIONS) + 2
    for r in range(next_row, next_row + 8):
        for c in range(1, 5):
            cell = ws6.cell(row=r, column=c)
            cell.fill = INPUT_FILL
            cell.border = BORDER
            cell.font = Font(name=FONT_NAME, size=10)
            if c == 4:
                dv_slot.add(cell)

    # ---------------- Sheet: Slot Reference ----------------
    ws4 = wb.create_sheet("Slot Reference")
    ws4.sheet_view.showGridLines = False
    for col, h in enumerate(["slot_id", "Canonical name", "Category"], start=1):
        cell = ws4.cell(row=1, column=col, value=h)
        style_header(cell)
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 38
    ws4.column_dimensions["C"].width = 16
    for r, (slot_id, slot_name, cat) in enumerate(SLOTS, start=2):
        ws4.cell(row=r, column=1, value=slot_id).font = Font(name=FONT_NAME, size=10)
        ws4.cell(row=r, column=2, value=slot_name).font = Font(name=FONT_NAME, size=10)
        ws4.cell(row=r, column=3, value=cat).font = Font(name=FONT_NAME, size=10)
        for c in range(1, 4):
            cell = ws4.cell(row=r, column=c)
            cell.fill = LOCKED_FILL
            cell.border = BORDER
    ws4.protection.sheet = False  # left unlocked but styled gray to signal "reference only"

    wb.save("worksheet.xlsx")
    print("Wrote worksheet.xlsx")


if __name__ == "__main__":
    build()
