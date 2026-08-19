"""
convert_worksheet.py

Reads the filled-in worksheet.xlsx (Requirements Matrix + CCC Catalog tabs)
and writes data/matched/*.json files in the exact same shape match_slots.py
would have produced -- so build_dist.py needs zero changes to consume this.

This SKIPS scrape_raw.py and match_slots.py entirely: because a human typed
the slot assignment directly (by checking a box in a known column), there's
nothing left to auto-match or classify. Every row this script writes is
articulation_kind='clean' and slot_confidence='manual'.

Usage:
    python convert_worksheet.py [--worksheet worksheet.xlsx] [--out-dir ../backend/data/matched]
"""

import argparse
import json
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent


def load_requirements(ws):
    """Returns list of {school_key, major_id, major_label, slots: [slot_id,...]}"""
    rows = list(ws.iter_rows(min_row=2, values_only=False))
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    slot_start_col = 5  # 0-indexed position of first slot column (col F = index 5)
    notes_col = len(header) - 1
    slot_headers = header[slot_start_col:notes_col]

    # Map header label -> slot_id, using Slot Reference sheet's name<->id pairing
    # (built by caller and passed in via name_to_id)
    results = []
    for row in rows:
        vals = [c.value for c in row]
        school_key, school_name, major_id, major_label, source_note = vals[:5]
        if not school_key or not major_id:
            continue
        slot_marks = vals[slot_start_col:notes_col]
        results.append({
            "school_key": str(school_key).strip(),
            "major_id": str(major_id).strip(),
            "major_label": major_label,
            "marks": dict(zip(slot_headers, slot_marks)),
        })
    return results, slot_headers


def load_slot_name_to_id(ws):
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        slot_id, name, cat = row[:3]
        if slot_id:
            mapping[name] = slot_id
    return mapping


def load_catalog(ws, ccc_labels):
    """Returns { ccc_key: {slot_id: course_code} }"""
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ccc_cols = {}  # ccc_key -> column index (0-based)
    label_to_key = {v: k for k, v in ccc_labels.items()}
    for i, h in enumerate(header):
        if h in label_to_key:
            ccc_cols[label_to_key[h]] = i

    catalog = {ccc: {} for ccc in ccc_labels}
    for row in ws.iter_rows(min_row=2, values_only=True):
        slot_id = row[0]
        if not slot_id:
            continue
        for ccc_key, col_idx in ccc_cols.items():
            code = row[col_idx]
            if code and str(code).strip():
                catalog[ccc_key][slot_id] = str(code).strip()
    return catalog


def load_electives(wb):
    """
    Returns {group_id: {school_key, major_id, pick_n, description,
    options: [{option_id, option_label, slots: [slot_id,...]}]}}

    NOTE: this data is captured and written to electives.json for future use,
    but build_dist.py does not currently enforce "pick N of M" logic -- the
    app's data model only understands flat required-slot lists today.
    """
    groups = {}
    if "Elective Groups" in wb.sheetnames:
        for row in wb["Elective Groups"].iter_rows(min_row=2, values_only=True):
            group_id, school_key, major_id, pick_n, description = (list(row) + [None] * 5)[:5]
            if not group_id:
                continue
            groups[str(group_id).strip()] = {
                "school_key": school_key, "major_id": major_id,
                "pick_n": pick_n, "description": description, "options": {},
            }

    if "Elective Options" in wb.sheetnames:
        for row in wb["Elective Options"].iter_rows(min_row=2, values_only=True):
            group_id, option_id, option_label, slot_id = (list(row) + [None] * 4)[:4]
            if not group_id or group_id not in groups:
                continue
            opts = groups[group_id]["options"]
            opts.setdefault(option_id, {"option_id": option_id, "option_label": option_label, "slots": []})
            if slot_id and str(slot_id).strip():
                opts[option_id]["slots"].append(str(slot_id).strip())

    # flatten options dict -> list per group
    for g in groups.values():
        g["options"] = list(g["options"].values())
    return groups


CCC_LABELS = {
    "deanza": "De Anza College", "evc": "Evergreen Valley College",
}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--worksheet", default=str(HERE / "worksheet.xlsx"))
    parser.add_argument("--out-dir", default=str(HERE / "backend" / "data" / "matched"))
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.worksheet, data_only=True)
    req_rows, slot_headers = load_requirements(wb["Requirements Matrix"])
    name_to_id = load_slot_name_to_id(wb["Slot Reference"])
    catalog = load_catalog(wb["CCC Catalog"], CCC_LABELS)
    electives = load_electives(wb)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    written, skipped_no_catalog_rows = 0, 0

    for req in req_rows:
        required_slot_ids = [
            name_to_id[name] for name, mark in req["marks"].items()
            if mark and str(mark).strip().upper() == "X" and name in name_to_id
        ]
        if not required_slot_ids:
            print(f"  (skip) {req['school_key']}/{req['major_id']}: no slots marked yet")
            continue

        for ccc_key in CCC_LABELS:
            rows = []
            for slot_id in required_slot_ids:
                code = catalog.get(ccc_key, {}).get(slot_id)
                rows.append({
                    "ccc_key": ccc_key,
                    "school_key": req["school_key"],
                    "major_id": req["major_id"],
                    "receiving_code": "(manual entry)",
                    "receiving_title": [n for n, i in name_to_id.items() if i == slot_id][0],
                    "slot_id": slot_id,
                    "slot_confidence": "manual",
                    "articulation_kind": "clean" if code else "no_articulation",
                    "sending_courses": [code] if code else [],
                    "articulation_detail": "entered via worksheet.xlsx" if code else "required by major, but no CCC Catalog entry filled in yet",
                })
                # IMPORTANT: the row always exists, even with no code, because
                # "this slot is required by this major" is a fact about the
                # UNIVERSITY, independent of whether any given CCC has a
                # course for it yet. Earlier versions of this script omitted
                # the row entirely when no code was filled in, which silently
                # made requirements vanish from school_majors.json rather
                # than showing up as an honest gap -- exactly the kind of
                # error this pipeline is supposed to prevent.

            out = {
                "ccc_key": ccc_key,
                "school_key": req["school_key"],
                "major_id": req["major_id"],
                "major_label": req["major_label"],
                "rows": rows,
            }
            out_path = out_dir / f"{ccc_key}__{req['school_key']}__{req['major_id']}.json"
            out_path.write_text(json.dumps(out, indent=2))
            written += 1
            if not any(r["sending_courses"] for r in rows):
                skipped_no_catalog_rows += 1

    print(f"\nWrote {written} matched file(s) to {out_dir}")
    if skipped_no_catalog_rows:
        print(f"({skipped_no_catalog_rows} of those have zero resolved courses -- fill in CCC Catalog tab to populate them)")

    if electives:
        electives_path = out_dir.parent / "electives.json"
        electives_path.write_text(json.dumps(electives, indent=2))
        print(f"\nWrote {len(electives)} elective group(s) to {electives_path}")
        print("(captured for reference -- build_dist.py / the app do not yet enforce 'pick N of M' logic)")

    print("\nNow run: python pipeline/build_dist.py  (from inside your backend/ folder)")


if __name__ == "__main__":
    main()
